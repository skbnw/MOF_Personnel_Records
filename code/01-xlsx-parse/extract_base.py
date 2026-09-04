# -*- coding: utf-8 -*-
"""
v1.0 既存HTMLのDATA抽出 + 裏表紙xlsxによる欠落補充
- 追加: 2016/2020/2022裏表紙から大学・高校・読みと、HTMLに無い年度キャリアを補充
- 追加: 半角カナ/全角英数の統一、ポスト表記のクレンジング
- 継承: 260513 index.html の DATA（2385人、〜令和4年度）
- 検証: 2022黒表紙の令和4ポストとDATAの差分レポート
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

CODE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(CODE_DIR))
from lib.common import (  # noqa: E402
    XLSX_2016,
    XLSX_2020,
    XLSX_2022,
    canon_kanji,
    canon_yomi,
    clean_post,
    dump_json,
    load_html_data,
    nfkc,
    normalize_person,
    unify_display,
)


def is_empty_post(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    if set(s) <= set("-－—–_　 "):
        return True
    return False


def iter_ura_people(path: Path, sheet_name: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    year_cols = []
    for i, h in enumerate(header):
        if h and "年度" in str(h):
            year_cols.append((i, str(h).strip()))
    for row in rows:
        name = nfkc(str(row[0] or "")).strip()
        y = nfkc(str(row[1] or "")).strip()
        if not name or not y:
            continue
        yomi = unify_display(str(row[2] or "")).strip()
        univ = unify_display(str(row[3] or "")).strip() if len(row) > 3 else ""
        hs = unify_display(str(row[4] or "")).strip() if len(row) > 4 else ""
        posts = {}
        for idx, lab in year_cols:
            if idx < len(row) and not is_empty_post(row[idx]):
                posts[lab] = clean_post(str(row[idx]))
        yield {
            "n": name,
            "y": y,
            "r": yomi,
            "u": univ,
            "h": hs,
            "posts": posts,
            "key": canon_kanji(name),
        }
    wb.close()


def build_lookup(people: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for p in people:
        out.setdefault(p["key"], []).append(p)
    return out


def pick_fill(cands: list[dict], field: str) -> str:
    for c in cands:
        v = (c.get(field) or "").strip()
        if v:
            return v
    return ""


def main(out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    data, _html = load_html_data()
    dump_json(out_dir / "data_from_html.json", data, indent=None)

    people_2022 = list(iter_ura_people(XLSX_2022, "黒表紙_2022"))
    people_2020 = list(iter_ura_people(XLSX_2020, "裏表紙"))
    people_2016 = list(iter_ura_people(XLSX_2016, "裏表紙"))
    lu2022 = build_lookup(people_2022)
    lu2020 = build_lookup(people_2020)
    lu2016 = build_lookup(people_2016)

    fill_rows = []
    filled = {"u": 0, "h": 0, "r": 0, "career_years": 0}
    unmatched_xlsx = 0
    for d in data:
        key = canon_kanji(d.get("n") or "")
        cands = lu2022.get(key) or lu2020.get(key) or lu2016.get(key) or []
        if not cands:
            unmatched_xlsx += 1
            fill_rows.append(
                {
                    "id": d["id"],
                    "n": d["n"],
                    "matched": 0,
                    "u_before": d.get("u") or "",
                    "h_before": d.get("h") or "",
                    "r_before": d.get("r") or "",
                    "u_after": d.get("u") or "",
                    "h_after": d.get("h") or "",
                    "r_after": d.get("r") or "",
                    "source": "",
                }
            )
            continue
        src = "2022" if key in lu2022 else ("2020" if key in lu2020 else "2016")
        before = {k: d.get(k) or "" for k in ("u", "h", "r")}
        for field in ("u", "h", "r"):
            if not (d.get(field) or "").strip():
                val = pick_fill(cands, field)
                if val:
                    d[field] = val
                    filled[field] += 1
        # 欠落年度のみ追記（既存ポストは上書きしない）。2022を優先。
        src_person = (lu2022.get(key) or lu2020.get(key) or lu2016.get(key) or [None])[0]
        if src_person:
            existing = {nfkc(c.get("年度") or "") for c in d.get("career") or []}
            career = d.setdefault("career", [])
            for lab, post in src_person["posts"].items():
                if nfkc(lab) in existing:
                    continue
                career.append({"年度": lab, "ポスト": post})
                existing.add(nfkc(lab))
                filled["career_years"] += 1
        fill_rows.append(
            {
                "id": d["id"],
                "n": d["n"],
                "matched": 1,
                "u_before": before["u"],
                "h_before": before["h"],
                "r_before": before["r"],
                "u_after": d.get("u") or "",
                "h_after": d.get("h") or "",
                "r_after": d.get("r") or "",
                "source": src,
            }
        )

    # 令和4 突合
    r4_diffs = []
    r4_label_candidates = ["令和４年度", "令和4年度"]
    data_by_key = {canon_kanji(d["n"]): d for d in data}
    for p in people_2022:
        post = ""
        for lab in r4_label_candidates:
            if lab in p["posts"]:
                post = p["posts"][lab]
                break
        if not post:
            # 全角半角ゆれ
            for lab, v in p["posts"].items():
                if "令和" in lab and "4" in nfkc(lab):
                    post = v
                    break
        d = data_by_key.get(p["key"])
        if not d:
            continue
        html_r4 = ""
        for c in d.get("career") or []:
            if "令和4" in nfkc(c.get("年度") or "") or "令和４" in (c.get("年度") or ""):
                html_r4 = c.get("ポスト") or ""
                break
        if (post or "") != (html_r4 or ""):
            r4_diffs.append(
                {
                    "n": p["n"],
                    "xlsx_r4": post,
                    "html_r4": html_r4,
                }
            )

    for d in data:
        normalize_person(d)
        career = d.get("career") or []
        if career:
            d["p"] = career[0].get("ポスト") or d.get("p") or ""

    dump_json(out_dir / "data_filled.json", data, indent=None)
    dump_json(
        out_dir / "xlsx_stats.json",
        {
            "html_people": len(data),
            "xlsx_2022": len(people_2022),
            "xlsx_2020": len(people_2020),
            "xlsx_2016": len(people_2016),
            "html_unmatched_to_xlsx": unmatched_xlsx,
            "filled": filled,
            "r4_diff_count": len(r4_diffs),
        },
    )

    def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)

    write_csv(
        out_dir / "fill_report.csv",
        fill_rows,
        [
            "id",
            "n",
            "matched",
            "u_before",
            "h_before",
            "r_before",
            "u_after",
            "h_after",
            "r_after",
            "source",
        ],
    )
    write_csv(out_dir / "r4_xlsx_vs_html.csv", r4_diffs, ["n", "xlsx_r4", "html_r4"])
    print("people", len(data), "filled", filled, "r4_diffs", len(r4_diffs), "xlsx_miss", unmatched_xlsx)


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    main(Path(args.out))
